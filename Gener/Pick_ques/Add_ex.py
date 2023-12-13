import openpyxl
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# define the headers
headers = ["Question", "Choice A", "Choice B",  "Choice C",  "Choice D", "Answer", "Level"]
for i in range(len(headers)):
    ws.cell(row=1, column=i+1).value = headers[i]

# add data
data = []
Data = []
Data.append({"a[1]": "Co_ban","start": 1, "range": 21})
Data.append({"a[2]": "Thong_dung","start": 21, "range": 32})
Data.append({"a[3]": "Van_dung","start": 32, "range": 41})
Data.append({"a[4]": "Van_dung_cao","start": 41, "range": 51})

def read_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
        return content
        
for item in Data:
    if item.get("a[1]") == "Co_ban":
        x = item.get("range")
        start = item.get("start")
        level = "Co_ban"
    elif  item.get("a[2]") == "Thong_dung":
        x = item.get("range")
        start = item.get("start")
        level = "Thong_dung"
    elif  item.get("a[3]") == "Van_dung":
        x = item.get("range")
        start = item.get("start")
        level = "Van_dung"
    elif  item.get("a[4]") == "Van_dung_cao":
        x = item.get("range")
        start = item.get("start") 
        level = "Van_dung_cao"
    i = start
    for i in range(i,x):
        Question = read_file(f"DATA/{level}/questions_data/Câu{i}/question.md")
        A = read_file(f"DATA/{level}/questions_data/Câu{i}/Choice_A.md")
        B = read_file(f"DATA/{level}/questions_data/Câu{i}/Choice_B.md")
        C = read_file(f"DATA/{level}/questions_data/Câu{i}/Choice_C.md")
        D = read_file(f"DATA/{level}/questions_data/Câu{i}/Choice_D.md")
        Level = level
        Answer = read_file(f"Ans_data/Câu{i}/ans.md")
        data.append([Question,A,B,C,D,Answer,Level])    
for n in range(len(data)):
    for m in range(len(data[n])):
        ws.cell(row=n+2, column=m+1).value = data[n][m]
wb.save('test.xlsx')